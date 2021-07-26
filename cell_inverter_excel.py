import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('C:\\Users\\CFY\\Desktop\\Python\\Projects\\Inverter_OG.xlsx')
active_sheet = wb.active
active_sheet.title = 'OriginalData'

#setting min and max cell values
maxrow = active_sheet.max_row
minrow = active_sheet.min_row
maxcol = active_sheet.max_column
mincol = active_sheet.min_column

mincell = get_column_letter(mincol) + str(minrow)
maxcell = get_column_letter(maxcol) + str(maxrow)

# collecting row and columns coordinates
col_list = []
row_list = []

for rowOfCellObjects in active_sheet[mincell:maxcell]:
    for cellObj in rowOfCellObjects:
        row_list.append(cellObj.row)
        col_list.append(cellObj.column)


new_rowlist = col_list
new_collist = row_list


#Create new worksheet for results
new_sheet = wb.create_sheet('InvertedResults')

#whichever list is greater the loop will run and put the data into the new file based on that
if len(new_collist) > len(new_rowlist):
    for idx in range(len(new_collist)):
        new_sheet.cell(row=new_rowlist[idx], column=new_collist[idx]).value = active_sheet.cell(row=row_list[idx], column=col_list[idx]).value
else:
    for idx in range(len(new_rowlist)):
        new_sheet.cell(row=new_rowlist[idx], column=new_collist[idx]).value = active_sheet.cell(row=row_list[idx], column=col_list[idx]).value

#saving excel file - use your own directory
wb.save('C:\\Users\\CFY\\Desktop\\Python\\Projects\\Inverter_OG_FINAL.xlsx')

